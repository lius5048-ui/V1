#!/usr/bin/env python3
"""Walmart Product Crawler → Excel Report.

Usage:
    python main.py "Sony WH-1000XM5"              # Search + details + all reviews
    python main.py "https://walmart.com/ip/123"    # URL direct
    python main.py "keyword" --excel report.xlsx   # Custom output
"""

import sys
import json
import os
import re
import argparse
import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Add parent to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from tools.walmart_crawler import search_products, fetch_product_details, fetch_product_reviews
from tools.fetch_all_reviews import fetch_all_reviews
from config import deepseek_client

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")


def crawl(query: str) -> dict:
    """Crawl Walmart product data and all reviews. Returns structured dict."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"\n🔍 Searching: {query}")

    # Step 1: Search
    products = search_products(query, limit=20)
    if not products:
        raise RuntimeError("No products found")

    # Pick best match (skip refurbished/open box if possible)
    main = products[0]
    for p in products:
        t = p.get("title", "").lower()
        if "refurbished" not in t and "open box" not in t and "restored" not in t:
            main = p
            break

    pid = main["product_id"]
    print(f"📦 {main['title'][:70]} (ID: {pid})")

    # Step 2: Product details
    print("📋 Fetching details...")
    details = fetch_product_details(pid)

    # Step 3: All reviews (paginated)
    print("⭐ Fetching ALL reviews...")
    all_reviews = fetch_all_reviews(pid, max_pages=20)

    # Step 4: Related products
    brand = details.get("brand") or main.get("brand", "")
    kw = brand if brand else query.split()[0]
    print(f"🔗 Related: {kw}")
    related = search_products(kw, limit=15)
    related = [p for p in related if p["product_id"] != pid]
    seen = set()
    unique = []
    for p in related:
        n = p["title"].strip().lower()[:40]
        if n not in seen:
            seen.add(n)
            unique.append(p)
    related = sorted(unique, key=lambda x: x.get("rating") or 0, reverse=True)[:10]

    result = {
        "query": query,
        "crawl_time": datetime.now().isoformat(),
        "main_product": main,
        "details": details,
        "reviews": all_reviews,
        "related_products": related,
    }

    # Save raw data
    safe = re.sub(r'[\\/*?:"<>|]', "_", query)[:30]
    path = os.path.join(OUTPUT_DIR, f"crawl_{safe}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, default=str, indent=2)
    print(f"💾 Data saved")
    return result


def _analyze_review_topics(reviews: list[dict]) -> str:
    """Use DeepSeek to categorize negative reviews into themes."""
    if not reviews:
        return "No negative reviews found."

    texts = []
    for r in reviews[:15]:  # Limit to 15 for API cost
        texts.append(f"[Rating {r.get('rating','?')}★] {r.get('reviewTitle',r.get('title',''))}: {r.get('reviewText',r.get('body',''))[:200]}")

    prompt = (
        "Categorize these negative product reviews into themes. "
        "Themes to check: product quality/defects, sizing/fit, shipping/delivery, "
        "customer service, value for money, comfort, battery life, noise cancellation, "
        "durability, or other.\n\n"
        + "\n---\n".join(texts) +
        "\n\nOutput format: For each theme, list: Theme name, count, representative quote(s). "
        "Keep it concise."
    )

    try:
        resp = deepseek_client.chat.completions.create(
            model="deepseek-chat",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
            max_tokens=2048,
        )
        return resp.choices[0].message.content
    except Exception as e:
        return f"Analysis unavailable: {e}"


def generate_excel(data: dict, output_path: str):
    """Generate Excel workbook with all product and review data."""
    wb = openpyxl.Workbook()

    # ========== Styles ==========
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    title_font = Font(bold=True, size=14, color="2F5496")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    neg_fill = PatternFill("solid", fgColor="FFF2CC")
    pos_fill = PatternFill("solid", fgColor="E2EFDA")

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(wrap_text=True, horizontal="center")
            cell.border = thin_border

    def auto_width(ws, max_width=50):
        for col in ws.columns:
            mx = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), max_width))
            ws.column_dimensions[col_letter].width = mx + 3

    # =====================================================
    # Sheet 1: Product Info
    # =====================================================
    ws1 = wb.active
    ws1.title = "Product Info"
    d = data["details"]
    main_p = data["main_product"]

    ws1.cell(row=1, column=1, value="Product Information Report").font = title_font
    ws1.merge_cells("A1:B1")
    ws1.cell(row=2, column=1, value=f"Crawl Date: {data['crawl_time'][:10]}").font = Font(italic=True, color="666666")

    fields = [
        ("Product ID", d.get("product_id", main_p.get("product_id", ""))),
        ("Title", d.get("title", main_p.get("title", ""))),
        ("Brand", d.get("brand", main_p.get("brand", ""))),
        ("Price", f"${d.get('price', 'N/A')}"),
        ("Currency", d.get("currency", "USD")),
        ("Rating", f"{d.get('rating', 'N/A')} / 5"),
        ("Total Reviews", d.get("review_count", 0)),
        ("Availability", d.get("availability", "")),
        ("SKU/UPC", d.get("sku", "")),
        ("URL", d.get("url", main_p.get("url", ""))),
    ]

    for i, (k, v) in enumerate(fields):
        ws1.cell(row=4 + i, column=1, value=k).font = Font(bold=True)
        ws1.cell(row=4 + i, column=2, value=str(v)[:200]).alignment = wrap
        ws1.cell(row=4 + i, column=1).border = thin_border
        ws1.cell(row=4 + i, column=2).border = thin_border

    # Description
    desc_row = 4 + len(fields) + 1
    desc = d.get("description", "")
    ws1.cell(row=desc_row, column=1, value="Description").font = Font(bold=True)
    ws1.cell(row=desc_row + 1, column=1, value=str(desc)[:2000]).alignment = wrap
    ws1.merge_cells(f"A{desc_row+1}:B{desc_row+1}")
    ws1.row_dimensions[desc_row + 1].height = 80

    # Specs
    specs_row = desc_row + 3
    ws1.cell(row=specs_row, column=1, value="Specifications").font = Font(bold=True, size=12, color="2F5496")
    ws1.merge_cells(f"A{specs_row}:B{specs_row}")
    specs = d.get("specs", {})
    for i, (k, v) in enumerate(specs.items()):
        r = specs_row + 1 + i
        ws1.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws1.cell(row=r, column=2, value=str(v)[:200]).alignment = wrap
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2).border = thin_border

    # Images
    img_row = specs_row + 1 + len(specs) + 1
    ws1.cell(row=img_row, column=1, value="Images").font = Font(bold=True)
    for i, url in enumerate(d.get("images", [])[:5]):
        ws1.cell(row=img_row + 1 + i, column=2, value=url).alignment = wrap

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 70

    # =====================================================
    # Sheet 2: Related Products
    # =====================================================
    ws2 = wb.create_sheet("Related Products")
    headers2 = ["#", "Product Name", "Price", "Rating", "Reviews", "URL"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))

    for i, p in enumerate(data["related_products"], 1):
        ws2.cell(row=i + 1, column=1, value=i)
        ws2.cell(row=i + 1, column=2, value=str(p.get("title", ""))[:100]).alignment = wrap
        ws2.cell(row=i + 1, column=3, value=p.get("price"))
        ws2.cell(row=i + 1, column=4, value=p.get("rating"))
        ws2.cell(row=i + 1, column=5, value=p.get("review_count", 0))
        ws2.cell(row=i + 1, column=6, value=p.get("url", ""))
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=i + 1, column=c).border = thin_border

    auto_width(ws2)

    # =====================================================
    # Sheet 3: All Reviews
    # =====================================================
    ws3 = wb.create_sheet("All Reviews")
    headers3 = ["#", "Rating", "Title", "Review Text", "Date", "Helpful", "Verified", "User"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, len(headers3))

    all_r = data["reviews"].get("all_reviews", [])
    all_r_sorted = sorted(all_r, key=lambda x: (x.get("rating", 5), -(x.get("helpful_count", x.get("positiveFeedback", 0)) or 0)))

    for i, rv in enumerate(all_r_sorted, 1):
        title = rv.get("reviewTitle", rv.get("title", ""))
        body = rv.get("reviewText", rv.get("body", rv.get("text", "")))
        date = rv.get("reviewSubmissionTime", rv.get("date", ""))
        helpful = rv.get("helpful_count", rv.get("positiveFeedback", 0))
        verified = rv.get("verified_purchase", rv.get("verified", False))
        nickname = rv.get("userNickname", rv.get("user_nickname", ""))

        rating = rv.get("rating", 0)
        ws3.cell(row=i + 1, column=1, value=i)
        ws3.cell(row=i + 1, column=2, value=rating)
        ws3.cell(row=i + 1, column=3, value=str(title)[:150]).alignment = wrap
        ws3.cell(row=i + 1, column=4, value=str(body)[:1000]).alignment = wrap
        ws3.cell(row=i + 1, column=5, value=str(date)[:20])
        ws3.cell(row=i + 1, column=6, value=helpful)
        ws3.cell(row=i + 1, column=7, value="Yes" if verified else "No")
        ws3.cell(row=i + 1, column=8, value=str(nickname)[:30])

        fill = neg_fill if rating <= 3 else (pos_fill if rating >= 4 else None)
        for c in range(1, len(headers3) + 1):
            cell = ws3.cell(row=i + 1, column=c)
            cell.border = thin_border
            cell.alignment = wrap
            if fill:
                cell.fill = fill
        ws3.row_dimensions[i + 1].height = 40

    auto_width(ws3)
    ws3.column_dimensions["D"].width = 60

    # =====================================================
    # Sheet 4: Negative Reviews Detail
    # =====================================================
    ws4 = wb.create_sheet("Negative Reviews")
    headers4 = ["#", "Rating", "Title", "Full Review Text", "Date", "Helpful", "Verified"]
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header(ws4, 1, len(headers4))

    neg_r = data["reviews"].get("negative_reviews", [])
    for i, rv in enumerate(neg_r, 1):
        title = rv.get("reviewTitle", rv.get("title", ""))
        body = rv.get("reviewText", rv.get("body", rv.get("text", "")))
        date = rv.get("reviewSubmissionTime", rv.get("date", ""))
        helpful = rv.get("helpful_count", rv.get("positiveFeedback", 0))
        verified = rv.get("verified_purchase", rv.get("verified", False))

        ws4.cell(row=i + 1, column=1, value=i)
        ws4.cell(row=i + 1, column=2, value=rv.get("rating", 0))
        ws4.cell(row=i + 1, column=3, value=str(title)[:150]).alignment = wrap
        ws4.cell(row=i + 1, column=4, value=str(body)[:1000]).alignment = wrap
        ws4.cell(row=i + 1, column=5, value=str(date)[:20])
        ws4.cell(row=i + 1, column=6, value=helpful)
        ws4.cell(row=i + 1, column=7, value="Yes" if verified else "No")
        for c in range(1, len(headers4) + 1):
            cell = ws4.cell(row=i + 1, column=c)
            cell.border = thin_border
            cell.fill = neg_fill
            cell.alignment = wrap
        ws4.row_dimensions[i + 1].height = 60

    # Analysis section
    analysis_row = len(neg_r) + 4
    ws4.cell(row=analysis_row, column=1, value="Negative Topics Analysis").font = Font(bold=True, size=12, color="2F5496")
    ws4.merge_cells(f"A{analysis_row}:G{analysis_row}")

    # Use DeepSeek for analysis
    print("  🤖 Analyzing review topics (DeepSeek)...")
    analysis = _analyze_review_topics(neg_r)
    ws4.cell(row=analysis_row + 1, column=1, value=analysis).alignment = wrap
    ws4.merge_cells(f"A{analysis_row+1}:G{analysis_row+1}")
    ws4.row_dimensions[analysis_row + 1].height = 200

    auto_width(ws4)
    ws4.column_dimensions["D"].width = 70

    # =====================================================
    # Sheet 5: Review Summary
    # =====================================================
    ws5 = wb.create_sheet("Review Summary")

    # Title
    ws5.cell(row=1, column=1, value=f"Review Summary").font = title_font
    ws5.merge_cells("A1:D1")

    # Stats
    rv_data = data["reviews"]
    total = rv_data.get("total_reviews", 0)
    total_captured = len(rv_data.get("all_reviews", []))
    neg_count = len(rv_data.get("negative_reviews", []))
    pos_count = len(rv_data.get("positive_reviews", []))
    avg_rating = d.get("rating", "N/A")

    stats = [
        ("Total Reviews (per Walmart)", total),
        ("Reviews Captured", total_captured),
        ("Capture Rate", f"{total_captured / total * 100:.0f}%" if total else "N/A"),
        ("Average Rating", avg_rating),
        ("Negative Reviews (1-3★)", neg_count),
        ("Positive Reviews (4-5★)", pos_count),
    ]

    for i, (k, v) in enumerate(stats):
        ws5.cell(row=3 + i, column=1, value=k).font = Font(bold=True)
        ws5.cell(row=3 + i, column=2, value=str(v))
        ws5.cell(row=3 + i, column=1).border = thin_border
        ws5.cell(row=3 + i, column=2).border = thin_border

    # Rating distribution
    dist_row = 11
    ws5.cell(row=dist_row, column=1, value="Rating Distribution").font = Font(bold=True, size=12, color="2F5496")
    ws5.merge_cells(f"A{dist_row}:B{dist_row}")

    # Count from page metadata
    reviews_page = None
    try:
        from curl_cffi import requests as curl_req
        import re as _re
        s2 = curl_req.Session()
        s2.get("https://www.walmart.com", impersonate="safari17_0", timeout=15)
        rp = s2.get(f"https://www.walmart.com/reviews/product/{d.get('product_id','')}?page=1", impersonate="safari17_0", timeout=15)
        m = _re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', rp.text, re.DOTALL)
        if m:
            nd = json.loads(m.group(1))
            reviews_page = nd.get("props",{}).get("pageProps",{}).get("initialData",{}).get("data",{}).get("reviews",{})
    except:
        pass

    if reviews_page:
        dist_data = [
            ("5★", reviews_page.get("ratingValueFiveCount", 0)),
            ("4★", reviews_page.get("ratingValueFourCount", 0)),
            ("3★", reviews_page.get("ratingValueThreeCount", 0)),
            ("2★", reviews_page.get("ratingValueTwoCount", 0)),
            ("1★", reviews_page.get("ratingValueOneCount", 0)),
        ]
    else:
        dist = {5: 0, 4: 0, 3: 0, 2: 0, 1: 0}
        for r in rv_data.get("all_reviews", []):
            rt = int(r.get("rating", 0))
            if rt in dist:
                dist[rt] += 1
        dist_data = sorted(dist.items(), reverse=True)

    for i, (star, cnt) in enumerate(dist_data):
        r = dist_row + 1 + i
        ws5.cell(row=r, column=1, value=star).border = thin_border
        ws5.cell(row=r, column=2, value=cnt).border = thin_border
        ws5.cell(row=r, column=3, value="█" * min(cnt, 40)).font = Font(color="2F5496")

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Rating Distribution"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Rating"
    chart.style = 10
    vals = Reference(ws5, min_col=2, min_row=dist_row + 1, max_row=dist_row + len(dist_data))
    cats = Reference(ws5, min_col=1, min_row=dist_row + 1, max_row=dist_row + len(dist_data))
    chart.add_data(vals, titles_from_data=False)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 10
    ws5.add_chart(chart, f"A{dist_row + len(dist_data) + 3}")

    # Additional stats from page data
    if reviews_page:
        ar = dist_row + len(dist_data) + 20
        ws5.cell(row=ar, column=1, value="Review Metadata").font = Font(bold=True, size=12, color="2F5496")
        ws5.merge_cells(f"A{ar}:B{ar}")
        meta = [
            ("Average Rating", reviews_page.get("averageOverallRating", "")),
            ("Recommended %", reviews_page.get("recommendedPercentage", "")),
            ("Reviews with Text", reviews_page.get("reviewsWithTextCount", "")),
        ]
        for i, (k, v) in enumerate(meta):
            ws5.cell(row=ar + 1 + i, column=1, value=k).font = Font(bold=True)
            ws5.cell(row=ar + 1 + i, column=2, value=str(v))

    ws5.column_dimensions["A"].width = 25
    ws5.column_dimensions["B"].width = 15
    ws5.column_dimensions["C"].width = 50

    # ========== Save ==========
    wb.save(output_path)
    print(f"📊 Excel saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Walmart Crawler → Excel Report")
    parser.add_argument("query", help="Walmart product URL or search keyword")
    parser.add_argument("--excel", "-e", default="", help="Output Excel path")
    parser.add_argument("--skip", "-s", action="store_true", help="Skip crawl, use cached JSON")
    args = parser.parse_args()

    query = args.query.strip()

    if args.skip:
        safe = re.sub(r'[\\/*?:"<>|]', "_", query)[:30]
        json_path = os.path.join(OUTPUT_DIR, f"crawl_{safe}.json")
        if not os.path.exists(json_path):
            print(f"No cached data found at {json_path}")
            sys.exit(1)
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
        print(f"Loaded cached data from {json_path}")
    else:
        data = crawl(query)

    # Generate Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r'[\\/*?:"<>|]', "_", query)[:20]
    filename = args.excel or f"walmart_report_{safe}_{timestamp}.xlsx"

    # Save to Windows Desktop by default
    desktop = os.path.expanduser("~/Desktop")
    if not os.path.isdir(desktop):
        desktop = os.path.expanduser("~")
    output_path = os.path.join(desktop, filename) if not args.excel else args.excel

    generate_excel(data, output_path)

    # Summary
    d = data["details"]
    rv = data["reviews"]
    neg = len(rv.get("negative_reviews", []))
    pos = len(rv.get("positive_reviews", []))
    print(f"\n{'='*50}")
    print(f"  ✅ Done!")
    print(f"  Product: {d.get('title','')[:60]}")
    print(f"  Price: ${d.get('price','N/A')}")
    print(f"  Reviews: {neg} negative + {pos} positive")
    print(f"  Excel: {output_path}")
    print(f"{'='*50}")
    print(f"  Sheets: Product Info | Related Products | All Reviews | Negative Reviews | Review Summary")


if __name__ == "__main__":
    main()
